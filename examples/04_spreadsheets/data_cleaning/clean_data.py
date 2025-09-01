"""
Cleans a messy Excel sheet using pandas.

- Assumes header row is Excel row 3 (cell B3 contains "Month, period").
- All numeric values are millimetres (mm) — strip "mm" and keep the number.
- Split "Jan,2001-2019" into Month, Start Year, End Year.
- Export with columns sized to content and numbers shown as 0.000.
"""

import numpy as np
import pandas as pd
from openpyxl.utils import get_column_letter  # convert 1->"A", 2->"B", ...

INPUT_FILE  = "04_spreadsheets/data_cleaning/messy_data.xlsx"
OUTPUT_FILE = "04_spreadsheets/data_cleaning/tidy_data.xlsx"


def to_mm(series: pd.Series) -> pd.Series:
    """Convert strings like '3.1mm' or '1,234' to numeric millimetres."""
    s = series.astype(str)
    s = s.str.replace(r"[,\s]", "", regex=True)      # remove commas/spaces
    s = s.str.replace(r"[a-zA-Z]+", "", regex=True)  # strip letters like 'mm'
    s = s.replace({"": np.nan, "nan": np.nan})
    return pd.to_numeric(s, errors="coerce")


def apply_formatting(ws, df: pd.DataFrame) -> None:
    """
    Apply number format to numeric columns and set simple column widths.
    - Numeric columns: "0.000" (rounded, zero-padded to 3 decimals)
    - Width: based on longest of header/cell display text
    """
    numeric_cols = [c for c in df.columns if pd.api.types.is_numeric_dtype(df[c])]

    for col_idx, col_name in enumerate(df.columns, start=1):
        col_letter = get_column_letter(col_idx)

        # Number format (skip header row = 1)
        if col_name in numeric_cols:
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                                    min_col=col_idx, max_col=col_idx):
                row[0].number_format = "0.000"

        # Simple width calculation
        max_len = len(str(col_name))
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                                min_col=col_idx, max_col=col_idx):
            val = row[0].value
            if val is None:
                text = ""
            elif col_name in numeric_cols:
                # show how it will display with 3 decimals
                try:
                    text = f"{float(val):.3f}"
                except Exception:
                    text = str(val)
            else:
                text = str(val)
            if len(text) > max_len:
                max_len = len(text)

        ws.column_dimensions[col_letter].width = min(60, max_len + 2)


def main() -> None:

    # pandas uses zero-based indexing, so Excel row 3 == row index 2.
    HEADER_ROW_INDEX = 2  
 
    # 1) Read the Excel file as a pandas dataframe containing only strings
    # Normally, pandas assumes the first row contains column headers.
    # Our messy spreadsheet has the headers in row 3, so we set header=None 
    # because we need to read the headers manually.
    df = pd.read_excel(INPUT_FILE, sheet_name=0, header=None, dtype=str)

    # 2) Set column names from the header row, then keep only the rows below it
    df.columns = df.iloc[HEADER_ROW_INDEX].astype(str).str.strip()
    df = df.iloc[HEADER_ROW_INDEX + 1 :].copy()

     # 3) Drop any completely empty rows/columns now that headers are set
    df = df.dropna(axis=0, how="all").dropna(axis=1, how="all")

    # 4) Split "Month, period" -> Month + Period
    month_period = df["Month, period"].astype(str)
    parts = month_period.str.split(",", n=1, expand=True)
    month_only = parts[0].str.strip().str.title()
    period_only = parts[1].str.strip()
    df.insert(0, "Month", month_only)
    df.insert(1, "Period", period_only)
    df = df.drop(columns=["Month, period"])

    # 5) Split Period "2001-2019" -> Start Year + End Year (simple split on '-')
    year_parts = period_only.str.split("-", n=1, expand=True)
    start_year = pd.to_numeric(year_parts[0].str.strip(), errors="coerce")
    end_year   = pd.to_numeric(year_parts[1].str.strip(), errors="coerce")
    df.insert(1, "Start Year", start_year)
    df.insert(2, "End Year", end_year)
    df = df.drop(columns=["Period"])

    # 6) Convert rainfall columns to numbers and give explicit names
    if "Lake Victoria" in df.columns:
        df["Lake Victoria (mm)"] = to_mm(df["Lake Victoria"])
        df = df.drop(columns=["Lake Victoria"])
    if "Simiyu" in df.columns:
        df["Simiyu (mm)"] = to_mm(df["Simiyu"])
        df = df.drop(columns=["Simiyu"])

    # 7) Save with formatting and simple auto-width
    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:

        sheet = "Sheet1"
        df.to_excel(writer, index=False, sheet_name=sheet)
        ws = writer.sheets[sheet]
        apply_formatting(ws, df)

    print(f"Cleaned data written to: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
